"""Polished recruiter-report exports for ranked TalentLens candidates."""

from __future__ import annotations

import csv
from datetime import date
from pathlib import Path
import re
from statistics import mean
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


REPORT_COLUMNS = [
    "Rank",
    "Candidate ID",
    "Candidate Name",
    "Current Role",
    "Experience",
    "Final Score",
    "Technical Score",
    "Experience Score",
    "Semantic Score",
    "Cognitive Twin Score",
    "Hidden Gem Score",
    "Hidden Gem Flag",
    "Authenticity Risk",
    "Missing Skills",
    "Strengths",
    "Weaknesses",
    "Recommendation",
    "Why Ranked",
]

SCORING_WEIGHTS = [
    ("Technical Fit", "34%"),
    ("Experience", "20%"),
    ("Behavioral Signals", "20%"),
    ("Authenticity", "15%"),
    ("Career Consistency", "11%"),
    ("Semantic Retrieval Blend", "15% semantic / 85% hybrid scoring"),
]

UNSUPPORTED_PHRASES = {
    "phone is not verified": "limited contact evidence found",
    "email is not verified": "limited contact evidence found",
    "identity is not verified": "profile evidence is incomplete",
    "missing one or more trust verification signals": "profile evidence is incomplete",
    "profile text appears sparse or generic": "profile evidence is incomplete",
    "when the role benefits from": "for roles requiring",
}


def build_ranked_report_rows(ranked_candidates: Iterable[Any]) -> list[dict[str, Any]]:
    """Convert RankedCandidate objects into clean, fixed recruiter-report rows."""

    rows: list[dict[str, Any]] = []
    for index, candidate in enumerate(ranked_candidates, start=1):
        missing = _missing_skills(candidate)
        rows.append(
            {
                "Rank": index,
                "Candidate ID": candidate.candidate_id,
                "Candidate Name": _candidate_name(candidate),
                "Current Role": _candidate_role(candidate),
                "Experience": _experience_text(candidate),
                "Final Score": _score(candidate.final_score),
                "Technical Score": _score(candidate.technical_fit.score),
                "Experience Score": _score(candidate.experience.score),
                "Semantic Score": _score(candidate.semantic_similarity.score),
                "Cognitive Twin Score": _score(candidate.recruiter_cognitive_twin.score),
                "Hidden Gem Score": _score(candidate.hidden_gem.score),
                "Hidden Gem Flag": "Yes" if candidate.hidden_gem.metadata.get("hidden_gem_flag") else "No",
                "Authenticity Risk": _authenticity_risk(candidate),
                "Missing Skills": _format_missing_skills(missing, candidate.final_score),
                "Strengths": _bullet_text(_strengths(candidate, missing)),
                "Weaknesses": _bullet_text(_weaknesses(candidate, missing)),
                "Recommendation": _recommendation(candidate, missing),
                "Why Ranked": _why_ranked(candidate, index, missing),
            }
        )
    return rows


def write_ranked_report_csv(rows: list[dict[str, Any]], output_path: str | Path) -> None:
    """Write polished ranked report rows to CSV with the fixed report columns."""

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=REPORT_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def write_ranked_report_xlsx(rows: list[dict[str, Any]], output_path: str | Path) -> None:
    """Write a formatted recruiter workbook with Rankings and Summary sheets."""

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    rankings_sheet = workbook.active
    rankings_sheet.title = "TalentLens Rankings"
    _write_rankings_sheet(rankings_sheet, rows)
    _write_summary_sheet(workbook.create_sheet("Summary"), rows)
    workbook.save(output_path)


def _write_rankings_sheet(sheet: Any, rows: list[dict[str, Any]]) -> None:
    title = "TalentLens AI Ranked Candidate Output"
    sheet.append([title])
    sheet.append([])
    sheet.append(REPORT_COLUMNS)
    for row in rows:
        sheet.append([row.get(column, "") for column in REPORT_COLUMNS])

    max_col = len(REPORT_COLUMNS)
    max_row = len(rows) + 3
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    sheet.freeze_panes = "A4"
    sheet.auto_filter.ref = f"A3:{get_column_letter(max_col)}{max_row}"

    title_fill = PatternFill("solid", fgColor="141824")
    header_fill = PatternFill("solid", fgColor="6C4DFF")
    white_font = Font(color="FFFFFF", bold=True)
    body_font = Font(color="172033")
    border = _thin_border()

    title_cell = sheet["A1"]
    title_cell.font = Font(color="FFFFFF", bold=True, size=16)
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 28
    sheet.row_dimensions[2].height = 6
    sheet.row_dimensions[3].height = 26

    score_columns = {
        "Rank",
        "Final Score",
        "Technical Score",
        "Experience Score",
        "Semantic Score",
        "Cognitive Twin Score",
        "Hidden Gem Score",
        "Hidden Gem Flag",
    }
    final_score_index = REPORT_COLUMNS.index("Final Score") + 1
    hidden_gem_index = REPORT_COLUMNS.index("Hidden Gem Flag") + 1

    for column_index, column_name in enumerate(REPORT_COLUMNS, start=1):
        cell = sheet.cell(row=3, column=column_index)
        cell.fill = header_fill
        cell.font = white_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        width = _column_width(column_name, rows)
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    for row_index in range(4, max_row + 1):
        sheet.row_dimensions[row_index].height = 78
        fill = PatternFill("solid", fgColor="F4F6FB" if row_index % 2 else "FFFFFF")
        for column_index, column_name in enumerate(REPORT_COLUMNS, start=1):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.font = body_font
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if column_name in score_columns else "left",
                vertical="top",
                wrap_text=True,
            )
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.0"
            if column_index == hidden_gem_index:
                is_gem = str(cell.value).lower() == "yes"
                cell.fill = PatternFill("solid", fgColor="DDF8EA" if is_gem else "E7EAF0")
                cell.font = Font(color="047857" if is_gem else "5B6472", bold=True)

    if rows:
        final_range = f"{get_column_letter(final_score_index)}4:{get_column_letter(final_score_index)}{max_row}"
        sheet.conditional_formatting.add(
            final_range,
            ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color="FDE2E2",
                mid_type="num",
                mid_value=75,
                mid_color="FFF1C7",
                end_type="num",
                end_value=100,
                end_color="CFF7DF",
            ),
        )


def _write_summary_sheet(sheet: Any, rows: list[dict[str, Any]]) -> None:
    top = rows[:5]
    avg_score = round(mean([float(row["Final Score"]) for row in rows]), 1) if rows else 0.0
    hidden_gems = sum(1 for row in rows if row["Hidden Gem Flag"] == "Yes")
    summary_rows = [
        ["TalentLens AI Ranked Candidate Output - Summary"],
        [],
        ["Metric", "Value"],
        ["Total candidates ranked", len(rows)],
        ["Top candidate", top[0]["Candidate Name"] if top else ""],
        ["Average final score", avg_score],
        ["Number of hidden gems", hidden_gems],
        ["Export generated date", date.today().isoformat()],
        [],
        ["Top 5 candidates", "", "", ""],
        ["Rank", "Candidate Name", "Final Score", "Recommendation"],
        *[[row["Rank"], row["Candidate Name"], row["Final Score"], row["Recommendation"]] for row in top],
        [],
        ["Scoring weights used", ""],
        *SCORING_WEIGHTS,
    ]
    for row in summary_rows:
        sheet.append(row)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    sheet.merge_cells(start_row=10, start_column=1, end_row=10, end_column=4)
    weights_row = 13 + len(top)
    sheet.merge_cells(start_row=weights_row, start_column=1, end_row=weights_row, end_column=2)
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 34
    sheet.column_dimensions["C"].width = 16
    sheet.column_dimensions["D"].width = 64

    border = _thin_border()
    for row in sheet.iter_rows():
        for cell in row:
            is_section = cell.row in {1, 10, weights_row}
            is_header = cell.row in {3, 11}
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor="141824" if is_section else "6C4DFF" if is_header else "F4F6FB" if cell.row % 2 else "FFFFFF")
            cell.font = Font(
                bold=is_section or is_header,
                color="FFFFFF" if is_section or is_header else "172033",
                size=16 if cell.row == 1 else 11,
            )
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.0"
        sheet.row_dimensions[row[0].row].height = 30 if row[0].row == 1 else 42 if row[0].row >= 12 and row[0].row <= 16 else 24


def _candidate_name(candidate: Any) -> str:
    profile = candidate.raw_candidate.get("profile") if isinstance(candidate.raw_candidate.get("profile"), dict) else {}
    return str(profile.get("name") or profile.get("full_name") or candidate.candidate_id)


def _candidate_role(candidate: Any) -> str:
    profile = candidate.raw_candidate.get("profile") if isinstance(candidate.raw_candidate.get("profile"), dict) else {}
    return str(profile.get("headline") or profile.get("role") or "Candidate Profile")


def _experience_text(candidate: Any) -> str:
    years = candidate.experience.metadata.get("candidate_years")
    try:
        number = float(years)
        return f"{number:.1f} years"
    except (TypeError, ValueError):
        return ""


def _score(value: Any) -> float | str:
    try:
        return round(float(value), 1)
    except (TypeError, ValueError):
        return ""


def _missing_skills(candidate: Any) -> list[str]:
    values = candidate.technical_fit.metadata.get("missing_skills", [])
    return _clean_list(values)


def _matched_skills(candidate: Any) -> list[str]:
    values = candidate.technical_fit.metadata.get("matched_skills", [])
    return _clean_list(values)


def _authenticity_risk(candidate: Any) -> str:
    risk = str(candidate.authenticity.metadata.get("risk_level", "MEDIUM")).upper()
    if risk == "LOW":
        return "Low - profile evidence is well supported by available signals"
    if risk == "HIGH":
        return "High - some claims require recruiter validation"
    return "Medium - profile evidence is incomplete in some areas"


def _format_missing_skills(missing: list[str], final_score: float) -> str:
    if not missing:
        return "No major missing skills"
    prefix = "Minor gaps" if final_score >= 78 or len(missing) <= 2 else "Missing"
    return f"{prefix}: {_format_list(missing[:5])}"


def _strengths(candidate: Any, missing: list[str]) -> list[str]:
    matched = _matched_skills(candidate)
    strengths: list[str] = []
    if matched:
        strengths.append(f"Strong match on {_format_list(matched[:3])}")
    years = candidate.experience.metadata.get("candidate_years")
    try:
        years_float = float(years)
        if years_float >= 5:
            strengths.append(f"{years_float:.1f} years of experience supports senior role readiness")
        elif years_float > 0:
            strengths.append(f"{years_float:.1f} years of hands-on experience for focused recruiter review")
    except (TypeError, ValueError):
        pass
    if candidate.semantic_similarity.score >= 78:
        strengths.append("High semantic match to the active role requirements")
    if candidate.skill_cluster.evidence:
        strengths.append(f"Coherent skill ecosystem: {candidate.skill_cluster.evidence[0]}")
    if candidate.hidden_gem.metadata.get("hidden_gem_flag"):
        strengths.append("Hidden-gem signal suggests growth potential")
    if not strengths and candidate.strengths:
        strengths.extend(candidate.strengths[:3])
    return _unique([_sanitize_text(item) for item in strengths])[:3]


def _weaknesses(candidate: Any, missing: list[str]) -> list[str]:
    weaknesses: list[str] = []
    if missing:
        weaknesses.append(f"Limited evidence for {_format_list(missing[:3])}")
    if candidate.experience.gaps:
        weaknesses.append(_sanitize_text(candidate.experience.gaps[0]))
    if candidate.career_consistency.gaps:
        weaknesses.append(_sanitize_text(candidate.career_consistency.gaps[0]))
    if not weaknesses:
        weaknesses.append("No major concerns; validate applied depth during recruiter screen")
    return _unique(weaknesses)[:3]


def _recommendation(candidate: Any, missing: list[str]) -> str:
    final_score = float(candidate.final_score)
    risk = str(candidate.authenticity.metadata.get("risk_level", "MEDIUM")).upper()
    if final_score >= 88 and risk != "HIGH" and len(missing) <= 1:
        level = "Strong Shortlist"
    elif final_score >= 78 and risk != "HIGH":
        level = "Shortlist"
    elif final_score >= 66:
        level = "Consider"
    elif final_score >= 55:
        level = "Backup Candidate"
    else:
        level = "Not Recommended"

    evidence: list[str] = []
    matched = _matched_skills(candidate)
    if matched:
        evidence.append(f"{_format_list(matched[:3])} fit")
    if candidate.semantic_similarity.score >= 78:
        evidence.append("strong semantic match")
    if not evidence and missing:
        evidence.append(f"needs validation on {_format_list(missing[:2])}")
    if not evidence:
        evidence.append("useful adjacent profile evidence")
    return f"{level} - {_sentence_case(', '.join(evidence))}."


def _why_ranked(candidate: Any, rank: int, missing: list[str]) -> str:
    matched = _matched_skills(candidate)
    strongest = f"{_format_list(matched[:3])} skill alignment" if matched else _candidate_role(candidate)
    gap = (
        f"The main gap is {_format_list(missing[:3])}."
        if missing
        else "No major required-skill gap is visible from the ranking evidence."
    )
    rank_reason = (
        "That rank is justified by stronger fit evidence than most candidates in this result set."
        if rank <= 3
        else "That rank reflects useful fit signals with a few recruiter-validation gaps."
    )
    return _sanitize_text(
        f"Ranked #{rank} with a final score of {candidate.final_score:.1f} due to {strongest}. {gap} {rank_reason}"
    )


def _bullet_text(items: list[str]) -> str:
    return "\n".join(f"- {_sanitize_text(item)}" for item in items)


def _clean_list(values: Any) -> list[str]:
    if not isinstance(values, list):
        values = str(values or "").split(",")
    return [str(value).strip() for value in values if str(value).strip()]


def _format_list(values: list[str]) -> str:
    return ", ".join(value.title().replace("Ml", "ML").replace("Ai", "AI").replace("Sql", "SQL") for value in values)


def _unique(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        key = value.lower()
        if key not in seen:
            result.append(value)
            seen.add(key)
    return result


def _sentence_case(value: str) -> str:
    return value[:1].upper() + value[1:] if value else value


def _sanitize_text(value: str) -> str:
    text = str(value or "")
    for unsafe, replacement in UNSUPPORTED_PHRASES.items():
        text = re.sub(re.escape(unsafe), replacement, text, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", text).strip()


def _column_width(column: str, rows: list[dict[str, Any]]) -> int:
    values = [str(row.get(column, "")) for row in rows]
    max_len = max([len(column), *(len(value) for value in values)] or [len(column)])
    if column in {"Strengths", "Weaknesses", "Recommendation", "Why Ranked", "Missing Skills", "Authenticity Risk"}:
        return min(max(max_len + 2, 24), 52)
    return min(max(max_len + 2, 12), 24)


def _thin_border() -> Border:
    side = Side(style="thin", color="D9DDEA")
    return Border(top=side, right=side, bottom=side, left=side)
